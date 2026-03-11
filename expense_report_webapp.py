#!/usr/bin/env python3
'''
# Install dependencies
pip install fastapi uvicorn requests openpyxl python-multipart

# Set your Groq API key
export GROQ_API_KEY='your_api_key_here'

# Run the server
python3 expense_report_webapp.py

# Access the web app
# Default model: http://localhost:8000
# Custom model: http://localhost:8000?model=your-model-name
'''
import base64
import json
import os
import uuid
from pathlib import Path
from typing import List
import requests
from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import uvicorn

app = FastAPI(title="Expense Report Generator")

# Create directories for storing files
UPLOAD_DIR = Path("uploads")
OUTPUT_DIR = Path("outputs")
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# Mount output directory for serving files
app.mount("/outputs", StaticFiles(directory="outputs"), name="outputs")


def get_image_mime_type(filename):
    """Determine MIME type from file extension."""
    ext = Path(filename).suffix.lower()
    mime_types = {
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.gif': 'image/gif',
        '.webp': 'image/webp'
    }
    return mime_types.get(ext, 'image/jpeg')


def process_images_batch(images_data, model, api_key, batch_num):
    """Process a batch of up to 5 images with Groq API."""
    
    system_prompt = """create an expense report with the following columns:
serial no, document number (invoice number from attachments), date, description (like hotel, transportation, lunch/dinner/food), vendor name, bill provided (yes), amount. In case invoice number is not available in the attachment, use license plate number or PNR.
For each receipt/bill, add one row with above columns populated.
Generate output in json format."""
    
    user_prompt = "Please analyze the attached receipt images and extract expense information according to the system prompt."
    
    # Build message content with images
    content = [{"type": "text", "text": user_prompt}]
    
    for img_data, filename in images_data:
        mime_type = get_image_mime_type(filename)
        content.append({
            "type": "image_url",
            "image_url": {
                "url": f"data:{mime_type};base64,{img_data}"
            }
        })
    
    # Call Groq API
    url = "https://api.groq.com/openai/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": content}
        ],
        "temperature": 0.1,
        "max_tokens": 4096
    }
    
    print(f"Sending batch {batch_num} ({len(images_data)} images) to Groq API...")
    response = requests.post(url, headers=headers, json=payload)
    
    if response.status_code != 200:
        raise Exception(f"API request failed: {response.status_code} - {response.text}")
    
    result = response.json()
    ai_response = result['choices'][0]['message']['content']
    
    print(f"Received response for batch {batch_num}")
    
    # Parse JSON from response
    try:
        if '```json' in ai_response:
            ai_response = ai_response.split('```json')[1].split('```')[0].strip()
        elif '```' in ai_response:
            ai_response = ai_response.split('```')[1].split('```')[0].strip()
        
        batch_data = json.loads(ai_response)
        return batch_data
    except json.JSONDecodeError as e:
        print(f"Failed to parse JSON response for batch {batch_num}: {e}")
        print(f"Response content: {ai_response[:500]}")
        raise


def extract_expense_data(images_data, model):
    """Extract expense data from receipt images using Groq API (max 5 images per batch)."""
    
    api_key = os.environ.get('GROQ_API_KEY')
    if not api_key:
        raise ValueError("GROQ_API_KEY environment variable not set")
    
    # Split images into batches of 5
    batch_size = 5
    image_batches = [images_data[i:i + batch_size] for i in range(0, len(images_data), batch_size)]
    
    print(f"Processing {len(images_data)} images in {len(image_batches)} batch(es)...")
    
    # Process each batch and collect results
    all_expenses = []
    
    for batch_num, batch in enumerate(image_batches, start=1):
        batch_data = process_images_batch(batch, model, api_key, batch_num)
        
        # Extract expenses from batch response
        if 'expenses' in batch_data:
            all_expenses.extend(batch_data['expenses'])
        elif isinstance(batch_data, list):
            all_expenses.extend(batch_data)
    
    # Renumber serial numbers sequentially
    for idx, expense in enumerate(all_expenses, start=1):
        expense['serial_no'] = idx
    
    # Combine all expenses into single structure
    merged_data = {
        'expenses': all_expenses
    }
    
    return merged_data


def create_excel_report(expense_data, output_file):
    """Create Excel report from expense data."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Add employee information at the top
    ws['A1'] = 'Employee ID:'
    ws['B1'] = expense_data.get('employee_id', '')
    ws['A2'] = 'Employee Name:'
    ws['B2'] = expense_data.get('employee_name', '')
    ws['A3'] = 'Reporting Manager:'
    ws['B3'] = expense_data.get('reporting_manager', '')
    
    # Make employee info bold
    for cell in ['A1', 'A2', 'A3']:
        ws[cell].font = Font(bold=True)
    
    # Add spacing
    current_row = 5
    
    # Add column headers
    headers = ['Serial No', 'Document Number', 'Date', 'Description', 
               'Vendor Name', 'Bill Provided', 'Amount']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row += 1
    
    # Add expense items
    expenses = expense_data.get('expenses', [])
    first_data_row = current_row
    
    for expense in expenses:
        ws.cell(row=current_row, column=1, value=expense.get('serial_no', ''))
        ws.cell(row=current_row, column=2, value=expense.get('document_number', ''))
        ws.cell(row=current_row, column=3, value=expense.get('date', ''))
        ws.cell(row=current_row, column=4, value=expense.get('description', ''))
        ws.cell(row=current_row, column=5, value=expense.get('vendor_name', ''))
        ws.cell(row=current_row, column=6, value=expense.get('bill_provided', ''))
        
        # Handle amount
        amount = expense.get('amount', 0)
        if isinstance(amount, str):
            amount = amount.replace('$', '').replace('₹', '').replace(',', '').strip()
            try:
                amount = float(amount)
            except ValueError:
                amount = 0
        
        ws.cell(row=current_row, column=7, value=amount)
        current_row += 1
    
    # Add total row with Excel formula
    total_row = current_row
    ws.cell(row=total_row, column=6, value='TOTAL:').font = Font(bold=True)
    
    total_cell = ws.cell(row=total_row, column=7)
    total_cell.value = f'=SUM(G{first_data_row}:INDIRECT("G"&ROW()-1))'
    total_cell.font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Save workbook
    wb.save(output_file)
    print(f"Expense report saved to: {output_file}")


@app.get("/", response_class=HTMLResponse)
async def get_home():
    """Serve the main webpage."""
    html_content = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Expense Report Generator</title>
        <style>
            * {
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }
            
            body {
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                min-height: 100vh;
                display: flex;
                justify-content: center;
                align-items: center;
                padding: 20px;
            }
            
            .container {
                background: white;
                border-radius: 16px;
                box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
                padding: 40px;
                max-width: 600px;
                width: 100%;
            }
            
            h1 {
                color: #333;
                margin-bottom: 10px;
                font-size: 28px;
            }
            
            .subtitle {
                color: #666;
                margin-bottom: 30px;
                font-size: 14px;
            }
            
            .form-group {
                margin-bottom: 24px;
            }
            
            label {
                display: block;
                margin-bottom: 8px;
                color: #333;
                font-weight: 500;
                font-size: 14px;
            }
            
            input[type="text"],
            input[type="file"] {
                width: 100%;
                padding: 12px;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                font-size: 14px;
                transition: border-color 0.3s;
            }
            
            input[type="text"]:focus {
                outline: none;
                border-color: #667eea;
            }
            
            input[type="file"] {
                padding: 10px;
                cursor: pointer;
            }
            
            .file-info {
                margin-top: 8px;
                font-size: 12px;
                color: #666;
            }
            
            button {
                width: 100%;
                padding: 14px;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 16px;
                font-weight: 600;
                cursor: pointer;
                transition: transform 0.2s, box-shadow 0.2s;
            }
            
            button:hover {
                transform: translateY(-2px);
                box-shadow: 0 10px 20px rgba(102, 126, 234, 0.4);
            }
            
            button:active {
                transform: translateY(0);
            }
            
            button:disabled {
                background: #ccc;
                cursor: not-allowed;
                transform: none;
            }
            
            .status {
                margin-top: 20px;
                padding: 16px;
                border-radius: 8px;
                display: none;
            }
            
            .status.loading {
                display: block;
                background: #e3f2fd;
                color: #1976d2;
                border: 1px solid #90caf9;
            }
            
            .status.success {
                display: block;
                background: #e8f5e9;
                color: #2e7d32;
                border: 1px solid #81c784;
            }
            
            .status.error {
                display: block;
                background: #ffebee;
                color: #c62828;
                border: 1px solid #ef5350;
            }
            
            .download-link {
                display: inline-block;
                margin-top: 12px;
                padding: 10px 20px;
                background: #4caf50;
                color: white;
                text-decoration: none;
                border-radius: 6px;
                font-weight: 500;
                transition: background 0.3s;
            }
            
            .download-link:hover {
                background: #45a049;
            }
            
            .model-info {
                background: #f5f5f5;
                padding: 12px;
                border-radius: 6px;
                margin-bottom: 24px;
                font-size: 13px;
                color: #555;
            }
            
            .spinner {
                display: inline-block;
                width: 16px;
                height: 16px;
                border: 3px solid rgba(25, 118, 210, 0.3);
                border-top-color: #1976d2;
                border-radius: 50%;
                animation: spin 0.8s linear infinite;
                margin-right: 8px;
                vertical-align: middle;
            }
            
            @keyframes spin {
                to { transform: rotate(360deg); }
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📊 Expense Report Generator</h1>
            <p class="subtitle">Generate expense reports from receipt images using AI</p>
            
            <div class="model-info">
                <strong>Model:</strong> <span id="modelName"></span>
            </div>
            
            <form id="expenseForm" enctype="multipart/form-data">
                <div class="form-group">
                    <label for="employeeId">Employee ID *</label>
                    <input type="text" id="employeeId" name="employeeId" required>
                </div>
                
                <div class="form-group">
                    <label for="employeeName">Employee Name *</label>
                    <input type="text" id="employeeName" name="employeeName" required>
                </div>
                
                <div class="form-group">
                    <label for="managerName">Reporting Manager Name *</label>
                    <input type="text" id="managerName" name="managerName" required>
                </div>
                
                <div class="form-group">
                    <label for="images">Receipt Images * (Multiple files allowed)</label>
                    <input type="file" id="images" name="images" accept="image/*" multiple required>
                    <div class="file-info">Supported formats: JPG, PNG, GIF, WebP</div>
                </div>
                
                <button type="submit" id="submitBtn">Generate Report</button>
            </form>
            
            <div id="status" class="status"></div>
        </div>
        
        <script>
            // Get model from query parameter or use default
            const urlParams = new URLSearchParams(window.location.search);
            const model = urlParams.get('model') || 'meta-llama/llama-4-scout-17b-16e-instruct';
            document.getElementById('modelName').textContent = model;
            
            document.getElementById('expenseForm').addEventListener('submit', async (e) => {
                e.preventDefault();
                
                const formData = new FormData();
                const employeeId = document.getElementById('employeeId').value;
                const employeeName = document.getElementById('employeeName').value;
                const managerName = document.getElementById('managerName').value;
                const images = document.getElementById('images').files;
                
                if (images.length === 0) {
                    showStatus('error', 'Please select at least one image');
                    return;
                }
                
                // Append form data
                formData.append('employee_id', employeeId);
                formData.append('employee_name', employeeName);
                formData.append('manager_name', managerName);
                formData.append('model', model);
                
                for (let i = 0; i < images.length; i++) {
                    formData.append('images', images[i]);
                }
                
                // Show loading status
                showStatus('loading', '<span class="spinner"></span>Processing images and generating report... This may take a minute.');
                document.getElementById('submitBtn').disabled = true;
                
                try {
                    const response = await fetch('/generate-report', {
                        method: 'POST',
                        body: formData
                    });
                    
                    const result = await response.json();
                    
                    if (response.ok) {
                        showStatus('success', 
                            `Report generated successfully! <br>
                            <a href="${result.download_url}" class="download-link" download>📥 Download Report</a>`
                        );
                    } else {
                        showStatus('error', `Error: ${result.detail || 'Failed to generate report'}`);
                    }
                } catch (error) {
                    showStatus('error', `Network error: ${error.message}`);
                } finally {
                    document.getElementById('submitBtn').disabled = false;
                }
            });
            
            function showStatus(type, message) {
                const statusDiv = document.getElementById('status');
                statusDiv.className = `status ${type}`;
                statusDiv.innerHTML = message;
            }
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html_content)


@app.post("/generate-report")
async def generate_report(
    employee_id: str = Form(...),
    employee_name: str = Form(...),
    manager_name: str = Form(...),
    model: str = Form(default="meta-llama/llama-4-scout-17b-16e-instruct"),
    images: List[UploadFile] = File(...)
):
    """API endpoint to generate expense report from uploaded images."""
    
    if not images:
        raise HTTPException(status_code=400, detail="No images provided")
    
    try:
        # Read and encode images
        images_data = []
        for img in images:
            content = await img.read()
            base64_img = base64.b64encode(content).decode('utf-8')
            images_data.append((base64_img, img.filename))
        
        # Extract expense data using Groq API
        expense_data = extract_expense_data(images_data, model)
        
        # Add employee details
        expense_data['employee_id'] = employee_id
        expense_data['employee_name'] = employee_name
        expense_data['reporting_manager'] = manager_name
        
        # Generate unique filename
        report_id = str(uuid.uuid4())
        output_filename = f"expense_report_{report_id}.xlsx"
        output_path = OUTPUT_DIR / output_filename
        
        # Create Excel report
        create_excel_report(expense_data, str(output_path))
        
        # Return download URL
        download_url = f"/outputs/{output_filename}"
        
        return {
            "success": True,
            "download_url": download_url,
            "filename": output_filename
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
