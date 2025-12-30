#!/usr/bin/env python3
'''
# Install dependencies
pip install requests openpyxl

# Set your Groq API key
export GROQ_API_KEY='your_api_key_here'

Execute using following cmd line:
python3 expense_report_gen.py --images chennai_expenses/*.jpg --employee-id 50 --employee-name "Kiran Lonikar" --manager-name "Ashwin Amalapuram" --output expense_report1.xlsx
'''
import argparse
import base64
import json
import os
import sys
from pathlib import Path
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def encode_image(image_path):
    """Encode image to base64 string."""
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')


def get_image_mime_type(image_path):
    """Determine MIME type from file extension."""
    ext = Path(image_path).suffix.lower()
    mime_types = {
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.gif': 'image/gif',
        '.webp': 'image/webp'
    }
    return mime_types.get(ext, 'image/jpeg')


def process_images_batch(images_batch, model, api_key, batch_num):
    """Process a batch of up to 5 images with Groq API."""
    
    system_prompt = """create an expense report with the following columns:
serial no, document number (invoice number from attachments), date, description (like hotel, transportation, lunch/dinner/food), vendor name, bill provided (yes), amount. In case invoice number is not available in the attachment, use license plate number.
For each receipt/bill, add one row with above columns populated.
Generate output in json format."""
    
    user_prompt = "Please analyze the attached receipt images and extract expense information according to the system prompt."
    
    # Build message content with images
    content = [{"type": "text", "text": user_prompt}]
    
    for img_path in images_batch:
        base64_img = encode_image(img_path)
        mime_type = get_image_mime_type(img_path)
        content.append({
            "type": "image_url",
            "image_url": {
                "url": f"data:{mime_type};base64,{base64_img}"
            }
        })
    
    # Call Groq API
    url = "https://api.groq.com/openai/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": content}
        ],
        "temperature": 0.1,
        "max_tokens": 4096
    }
    
    print(f"Sending batch {batch_num} ({len(images_batch)} images) to Groq API...")
    response = requests.post(url, headers=headers, json=payload)
    
    if response.status_code != 200:
        raise Exception(f"API request failed: {response.status_code} - {response.text}")
    
    result = response.json()
    
    # Extract the response content
    ai_response = result['choices'][0]['message']['content']
    
    print(f"Received response for batch {batch_num}")
    
    # Try to parse JSON from the response
    try:
        # Remove markdown code blocks if present
        if '```json' in ai_response:
            ai_response = ai_response.split('```json')[1].split('```')[0].strip()
        elif '```' in ai_response:
            ai_response = ai_response.split('```')[1].split('```')[0].strip()
        
        batch_data = json.loads(ai_response)
        return batch_data
    except json.JSONDecodeError as e:
        print(f"Failed to parse JSON response for batch {batch_num}: {e}")
        print(f"Response content: {ai_response[:500]}")
        raise


def extract_expense_data(images, model):
    """Extract expense data from receipt images using Groq API (max 5 images per batch)."""
    
    api_key = os.environ.get('GROQ_API_KEY')
    if not api_key:
        raise ValueError("GROQ_API_KEY environment variable not set")
    
    # Split images into batches of 5
    batch_size = 5
    image_batches = [images[i:i + batch_size] for i in range(0, len(images), batch_size)]
    
    print(f"Processing {len(images)} images in {len(image_batches)} batch(es)...")
    
    # Process each batch and collect results
    all_expenses = []
    
    for batch_num, batch in enumerate(image_batches, start=1):
        batch_data = process_images_batch(batch, model, api_key, batch_num)
        
        # Extract expenses from batch response
        if 'expenses' in batch_data:
            all_expenses.extend(batch_data['expenses'])
        elif isinstance(batch_data, list):
            # Handle case where response is directly a list of expenses
            all_expenses.extend(batch_data)
    
    # Renumber serial numbers sequentially
    for idx, expense in enumerate(all_expenses, start=1):
        expense['serial_no'] = idx
    
    # Combine all expenses into single structure
    merged_data = {
        'expenses': all_expenses
    }
    
    return merged_data


def create_excel_report(expense_data, output_file):
    """Create Excel report from expense data."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Add employee information at the top
    ws['A1'] = 'Employee ID:'
    ws['B1'] = expense_data.get('employee_id', '')
    ws['A2'] = 'Employee Name:'
    ws['B2'] = expense_data.get('employee_name', '')
    ws['A3'] = 'Reporting Manager:'
    ws['B3'] = expense_data.get('reporting_manager', '')
    
    # Make employee info bold
    for cell in ['A1', 'A2', 'A3']:
        ws[cell].font = Font(bold=True)
    
    # Add spacing
    current_row = 5
    
    # Add column headers
    headers = ['Serial No', 'Document Number', 'Date', 'Description', 
               'Vendor Name', 'Bill Provided', 'Amount']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row += 1
    
    # Add expense items
    expenses = expense_data.get('expenses', [])
    first_data_row = current_row
    
    for expense in expenses:
        ws.cell(row=current_row, column=1, value=expense.get('serial_no', ''))
        ws.cell(row=current_row, column=2, value=expense.get('document_number', ''))
        ws.cell(row=current_row, column=3, value=expense.get('date', ''))
        ws.cell(row=current_row, column=4, value=expense.get('description', ''))
        ws.cell(row=current_row, column=5, value=expense.get('vendor_name', ''))
        ws.cell(row=current_row, column=6, value=expense.get('bill_provided', ''))
        
        # Handle amount - try to convert to float
        amount = expense.get('amount', 0)
        if isinstance(amount, str):
            # Remove currency symbols (USD, INR) and commas
            amount = amount.replace('$', '').replace('₹', '').replace(',', '').strip()
            try:
                amount = float(amount)
            except ValueError:
                amount = 0
        
        ws.cell(row=current_row, column=7, value=amount)
        current_row += 1
    
    last_data_row = current_row - 1
    
    # Add total row with Excel formula
    total_row = current_row
    ws.cell(row=total_row, column=6, value='TOTAL:').font = Font(bold=True)
    
    # Use dynamic formula that sums from first data row to the cell right above this total cell
    # INDIRECT with ROW() makes it truly dynamic - it always sums up to the row above
    total_cell = ws.cell(row=total_row, column=7)
    total_cell.value = f'=SUM(G{first_data_row}:INDIRECT("G"&ROW()-1))'
    total_cell.font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Save workbook
    wb.save(output_file)
    print(f"Expense report saved to: {output_file}")


def main():
    parser = argparse.ArgumentParser(
        description='Generate expense report from receipt images using Groq API'
    )
    
    parser.add_argument(
        '--images',
        nargs='+',
        required=True,
        help='Paths to receipt images'
    )
    
    parser.add_argument(
        '--employee-id',
        required=True,
        help='Employee ID'
    )
    
    parser.add_argument(
        '--employee-name',
        required=True,
        help='Employee name'
    )
    
    parser.add_argument(
        '--manager-name',
        required=True,
        help='Reporting manager name'
    )
    
    parser.add_argument(
        '--output',
        required=True,
        help='Output Excel file path'
    )
    
    parser.add_argument(
        '--model',
        default='meta-llama/llama-4-scout-17b-16e-instruct',
        help='Groq model to use (default: meta-llama/llama-4-scout-17b-16e-instruct)'
    )
    
    args = parser.parse_args()
    
    # Validate image files exist
    for img_path in args.images:
        if not os.path.exists(img_path):
            print(f"Error: Image file not found: {img_path}")
            sys.exit(1)
    
    try:
        # Extract expense data using Groq API (only images)
        expense_data = extract_expense_data(
            args.images,
            args.model
        )
        
        # Add employee details to the expense data
        expense_data['employee_id'] = args.employee_id
        expense_data['employee_name'] = args.employee_name
        expense_data['reporting_manager'] = args.manager_name
        
        # Create Excel report
        create_excel_report(expense_data, args.output)
        
        print("Expense report generated successfully!")
        
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
